# 好的，您可以使用Python中的Selenium库来模拟浏览器访问网站并复制所有内容，以下是代码示例：
# pythonCopy Code

import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

# 引入必要的库
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import re
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl


# username: sntxq password: 462167ysy


# 全局变量


def file_line_generator(file_path):
    """
    生成器函数，用于逐行读取文件内容。

    参数:
    file_path (str): 文件的路径。

    返回:
    generator: 一个生成器，每次迭代返回文件的下一行。
    """
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            yield line.strip()  # 使用 strip() 去除每行的首尾空白字符


def get_driver():
    try:
        # 设置浏览器驱动路径（这里使用Chrome浏览器驱动作为示例）
        driver_path = 'E:\study\coding\python\crawler\chromedriver\chromedriver-win64\chromedriver.exe'

        # 选择谷歌浏览器
        s = Service(executable_path=driver_path)
        driver = webdriver.Chrome(service=s)
        return driver
    except Exception as e:
        print(f'获取浏览器实例失败！')
        # print(f'错误是：{e}')
        return webdriver.Chrome()


def get_all_text(element):
    """
    递归获取元素及其所有子元素的文本内容。

    参数:
    driver: WebDriver实例。
    element: 要获取文本的元素。

    返回:
    str: 元素及其所有子元素的文本内容。
    """
    all_text = element.text + "\n"  # 获取当前元素的文本并添加换行符

    # # 获取当前元素的所有子元素
    # for child in element.find_elements(By.XPATH, './*'):
    #     all_text += get_all_text(driver, child)  # 递归获取每个子元素的文本

    return all_text


# 爬取某些设置下的基础内容
def get_crawler():
    # 使用示例
    # 假设你有一个名为 'example.txt' 的文件

    file = file_line_generator('AKC_breed_url.txt')

    traits = {}
    traits_terms = ['AKC_breed']

    for each in file:
        parts = each.strip('\n').split('\t')
        breed = parts[0]
        print(f'正在处理：{breed}')
        # 初始化
        traits[breed] = {}
        traits[breed]['AKC_breed'] = breed
        url = parts[1]

        driver.get(url)
        # 等待页面加载完毕，不然有可能找不到元素
        time.sleep(5)

        # 基本统计量
        try:
            statistics = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div[2]')
            # 获取目标元素下的所有文本
            all_text = get_all_text(statistics)
        except Exception as e:
            all_text = str(e)
        with open(f'./AKC_results/{breed.replace(" ", "_")}_statistics.txt', 'w') as f:
            f.write(all_text)
        f.close()
        # print(f'基本统计量为：{all_text}')

        # 表型按钮
        trait_element = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div[3]/div/div[1]/ul/li[5]")
        # # 使用JavaScript代码滚动到元素的位置
        # driver.execute_script("arguments[0].scrollIntoView();", trait_element)
        trait_element.click()

        children_elements = driver.find_element(By.XPATH, '//*[@id="breed-page__traits__all"]/div/div').find_elements(
            By.XPATH, "./*")

        # 遍历所有子元素，并获取它们的属性值，这里以'data-attribute'为例
        for child in children_elements:
            header = child.find_element(By.CSS_SELECTOR, 'h4').text
            if header in traits_terms:
                pass
            else:
                traits_terms.append(header)

            try:
                choice_ele = child.find_element(By.CLASS_NAME, 'breed-trait-score').find_element(By.CLASS_NAME,
                                                                                                 'breed-trait-score__choices')

                # 获取该元素下所有的子元素
                all_choice = choice_ele.find_elements(By.CSS_SELECTOR, "div")

                # 遍历所有子元素，获取它们的class属性
                selected_text = ''
                for each_star in all_choice:
                    class_attr = each_star.get_attribute('class')
                    if 'selected' in class_attr:
                        selected_text += each_star.text + '-'
                selected_text = selected_text.strip('-')
                print(f'{header}:{selected_text}')

                traits[breed][header] = selected_text

            except  Exception as e:
                # print(f'错误信息是:{e}')
                star_ele = child.find_element(By.CLASS_NAME, 'breed-trait-score').find_element(By.CLASS_NAME,
                                                                                               'breed-trait-score__score-wrap')
                # 获取该元素下所有的子元素
                all_star = star_ele.find_elements(By.TAG_NAME, "*")

                selected_stars_num = 0
                # 遍历所有子元素，获取它们的class属性
                for each_star in all_star:
                    class_attr = each_star.get_attribute('class')
                    if 'filled' in class_attr:
                        selected_stars_num += 1
                print(f'{header}:{selected_stars_num}')

                traits[breed][header] = selected_stars_num

    # sig = True
    # file = file_line_generator('tree_sort_test.txt')
    # for line in file:
    #     breed = line.strip('\n')
    #
    #     try:
    #         driver.get(url)
    #         # 等待页面加载完毕，不然有可能找不到元素
    #         time.sleep(5)
    #
    #         # 先刷新input 文本框的内容
    #         input_element = driver.find_element(By.XPATH,
    #                                             '//*[@id="form-02f15b39-5275-4c9e-ae23-2fe8460c481e"]/div[2]/div[1]/input')  # 获取该输入框的ID
    #
    #         # # 创建 ActionChains 对象
    #         # action_chains = ActionChains(driver)
    #         #
    #         # # 移动到目标元素的位置
    #         # action_chains.move_to_element(input_element).perform()
    #
    #         js = "window.scrollTo(0,document.body.scrollHeight)"
    #         driver.execute_script(js)
    #
    #         time.sleep(3)
    #
    #         input_element.clear()  # 清除该输入框中的原本内容
    #         input_element.send_keys(breed, Keys.ENTER)  # 向该输入框中添加要查询的关键字
    #         time.sleep(3)
    #
    #         result_link = driver.find_element(By.XPATH,
    #                                           '//*[@id="block-02f15b39-5275-4c9e-ae23-2fe8460c481e"]/div[1]/article[1]/figure/a')
    #         result_link.click()  # 可能会报错，原因可能是由悬浮元素遮挡
    #
    #         time.sleep(3)
    #         js = "window.scrollTo(0,document.body.scrollHeight)"
    #         driver.execute_script(js)
    #         time.sleep(3)
    #
    #         # breed品种
    #         query_breed = breed
    #         result_breed = driver.find_element(By.XPATH, ' //*[@id="primary"]/div[1]/div[1]/h1').text
    #         # print(f'查询品种为：{query_breed}, 当前品种为：{result_breed}')
    #
    #         # 基本统计量
    #         statistics = driver.find_element(By.XPATH, '//*[@id="primary"]/div[1]/div[2]/section')
    #         # 获取目标元素下的所有文本
    #         all_text = get_all_text(driver, statistics)
    #         with open(f'./dogtime_results/{result_breed.replace(" ", "_")}_statistics.txt', 'w') as f:
    #             f.write(all_text)
    #         f.close()
    #         # print(f'基本统计量为：{all_text}')
    #
    #         # quick facts
    #         facts = driver.find_element(By.XPATH, '//*[@id="primary"]/div[1]/div[1]/div[2]/ul')
    #         all_text = get_all_text(driver, facts)
    #         with open(f'./dogtime_results/{result_breed.replace(" ", "_")}_statistics.txt', 'a+') as f:
    #             f.write('\n' + all_text)
    #         f.close()
    #
    #         # 行为表型
    #         current_behavior = []
    #         for i in range(len(behavior_terms)):
    #             current_behavior.append('null')
    #         current_behavior[0] = query_breed
    #         current_behavior[1] = result_breed
    #
    #         # 找到包含特定class的元素的父元素
    #         parent_element = driver.find_element(By.CLASS_NAME, 'xe-breed-card')
    #
    #         # 在父元素下找到所有具有特定class的子元素
    #         elements = parent_element.find_elements(By.CLASS_NAME, 'xe-breed-accordion')
    #
    #         # #test
    #         # test = driver.find_element(By.XPATH, '//*[@id="primary"]/div[1]/div[1]/div[2]/div[12]/details[1]/ul/li[1]/details/summary/span[2]/span[1]')
    #         # element_class = test.get_attribute("class")
    #         # print(element_class)
    #
    #         if sig == True:
    #             # 遍历所有找到的元素
    #             for element in elements:
    #                 # 找到summary heading
    #                 summary_head = element.find_element(By.CLASS_NAME, 'xe-breed-accordion__summary-heading')
    #                 summary_text = summary_head.text
    #                 behavior_terms.append(summary_text)
    #
    #                 # 找到下拉按钮并且点击
    #                 pulldown = element.find_element(By.CLASS_NAME, 'xe-expander')
    #                 # 使用JavaScript代码滚动到元素的位置
    #                 driver.execute_script("arguments[0].scrollIntoView();", pulldown)
    #                 pulldown.click()
    #
    #                 sub_elements = element.find_elements(By.CLASS_NAME, 'xe-breed-characteristics-list__heading')
    #                 for sub in sub_elements:
    #                     sub_text = sub.text
    #                     behavior_terms.append(sub_text)
    #
    #             # 初始化完成
    #             behavior.append(behavior_terms)
    #             print(f'初始化后的行为共有：{behavior_terms}')
    #
    #         if sig == False:
    #
    #             # 遍历所有找到的元素
    #             # count = 0
    #             for element in elements:
    #                 # 找到summary heading
    #                 summary_head = element.find_element(By.CLASS_NAME, 'xe-breed-accordion__summary-heading')
    #                 summary_text = summary_head.text
    #                 # print(summary_text)
    #
    #                 # 找到下拉按钮并且点击
    #                 pulldown = element.find_element(By.CLASS_NAME, 'xe-expander')
    #                 # 使用JavaScript代码滚动到元素的位置
    #                 driver.execute_script("arguments[0].scrollIntoView();", pulldown)
    #                 pulldown.click()
    #
    #                 # 找到summary对应的评分
    #                 summary_stars = element.find_element(By.CLASS_NAME, 'xe-breed-accordion__summary').find_element(
    #                     By.CLASS_NAME, 'xe-breed-star-rating')
    #
    #                 # 获取该元素下所有的子元素
    #                 all_elements = summary_stars.find_elements(By.TAG_NAME, "*")
    #
    #                 summary_selected_stars_num = 0
    #                 # 遍历所有子元素，获取它们的class属性
    #                 for each_star in all_elements:
    #                     class_attr = each_star.get_attribute('class')
    #                     if 'selected' in class_attr:
    #                         summary_selected_stars_num += 1
    #                 # summary_selected_stars = summary_stars.find_elements(By.CLASS_NAME,
    #                 #                                                      'xe-breed-star xe-breed-star--selected')
    #                 # summary_selected_stars_num = len(summary_selected_stars)
    #                 # print(f'{summary_text}：{summary_selected_stars_num}')
    #
    #                 # element_class = summary_stars.get_attribute("class")
    #                 #
    #                 # # 打印获取到的class属性
    #                 # print(element_class)
    #
    #                 if summary_text in behavior_terms:
    #                     index = behavior_terms.index(summary_text)
    #                     current_behavior[index] = summary_selected_stars_num
    #                 else:
    #                     print(f'出错了！现在的summary_text:"{summary_text}"')
    #                     # current_behavior.append('null')
    #
    #                 sub_elements = element.find_elements(By.CLASS_NAME, 'xe-breed-characteristics-list__heading')
    #                 for sub in sub_elements:
    #                     sub_text = sub.text
    #                     sub_stars = sub.find_element(By.XPATH, 'following-sibling::*[1]')
    #
    #                     # 获取该元素下所有的子元素
    #                     all_elements = sub_stars.find_elements(By.TAG_NAME, "*")
    #
    #                     sub_selected_stars_num = 0
    #                     # 遍历所有子元素，获取它们的class属性
    #                     for each_star in all_elements:
    #                         class_attr = each_star.get_attribute('class')
    #                         if 'selected' in class_attr:
    #                             sub_selected_stars_num += 1
    #
    #                     # print(f'{sub_text}：{sub_selected_stars_num}')
    #
    #                     # sub_selected_stars = sub_stars.find_elements(By.CLASS_NAME,
    #                     #                                              'xe-breed-star xe-breed-star--selected')
    #                     # sub_selected_stars_num = len(sub_selected_stars)
    #                     # print(f'{sub_text}：{sub_selected_stars_num}')
    #
    #                     if sub_text in behavior_terms:
    #                         index = behavior_terms.index(sub_text)
    #                         current_behavior[index] = sub_selected_stars_num
    #                     else:
    #                         print(f'出错了！现在的sub_text:"{sub_text}"')
    #                         # current_behavior.append('null')
    #
    #         behavior.append(current_behavior)
    #         print(f'当前行为表型为:{current_behavior}')
    #
    #         sig = False
    #
    #     except Exception as e:
    #         with open('error_info.txt', 'a+') as fr:
    #             fr.write(f'出错样本是:{breed}, 出错原因是:{e}' + '\n')
    #
    # export_info(behavior)
    export_info(traits, traits_terms)


def get_breed_url(url):
    driver.get(url)

    # 定位到父元素，这里以一个假设的XPath为例
    parent_element = driver.find_element(By.XPATH, "/html/body/div[4]/div[1]/div[3]/aside/div/div[2]/div/div")
    parent_element.click()

    time.sleep(5)

    # test = driver.find_element(By.XPATH, "/html/body/div[4]/div[1]/div[3]/aside/div/div[2]/div/div/div[2]/div")
    # attri = test.get_attribute('class')
    # print(attri)

    # 找到父元素下一级的所有子元素，这里以class为'child'的元素为例
    # children_elements = driver.find_element(By.CLASS_NAME, "selectize-dropdown-content").find_elements(By.XPATH, "./*")
    children_elements = driver.find_element(By.XPATH,
                                            "/html/body/div[4]/div[1]/div[3]/aside/div/div[2]/div/div/div[2]/div").find_elements(
        By.XPATH, "./*")

    # 遍历所有子元素，并获取它们的属性值，这里以'data-attribute'为例
    with open('AKC_breed_url.txt', 'w') as f:
        for child in children_elements:
            # breed = child.text  # 这里获取文本失败估计是元素不可见
            attribute_value = child.get_attribute('data-value')
            breed = attribute_value.strip('/').split('/')[-1]
            print(f'{breed}:{attribute_value}')
            f.write(breed + '\t' + attribute_value + '\n')


def export_info(trait, trait_terms):
    outlist = []
    outlist.append(trait_terms)
    for key, value in trait.items():
        current = []
        for each_trait in trait_terms:
            if each_trait in value:
                current.append(str(value[each_trait]))
            else:
                current.append('null')
        outlist.append(current)

    sheetname = "AKC"

    # 创建输出表格Excel：创建工作表
    excel = openpyxl.Workbook()
    # 创建sheet页：以demo为名字创建一个sheet页
    sheet = excel.create_sheet(f'{sheetname}', 0)

    # 第一行第一列的单元格
    for i in range(len(outlist)):
        for k in range(len(outlist[i])):
            cell = sheet.cell(row=i + 1, column=k + 1)
            # 单元格赋值
            cell.value = outlist[i][k]

    # 保存excel文件
    excel.save(f'AKC_alltraits_info.xlsx')


if __name__ == "__main__":
    # 设置你想要访问的基础网址
    web_url = 'https://www.akc.org/dog-breeds/'
    driver = get_driver()

    # get_breed_url(web_url)

    get_crawler()
    # export_info()

    # 休息一分钟，检查信息
    time.sleep(5)

    # 不关闭浏览器
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("t").key_up(Keys.CONTROL).perform()
