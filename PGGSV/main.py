# 好的，您可以使用Python中的Selenium库来模拟浏览器访问网站并复制所有内容，以下是代码示例：
# pythonCopy Code

import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

# 引入必要的库
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import re
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import csv

# 全局变量
sv_info = []
error_sv_info = []


def get_driver():
    try:
        # 设置浏览器驱动路径（这里使用Chrome浏览器驱动作为示例）
        driver_path = 'E:\study\coding\python\crawler\chromedriver\chromedriver-win64\chromedriver.exe'

        # 选择谷歌浏览器
        s = Service(executable_path=driver_path)
        driver = webdriver.Chrome(service=s)
        return driver
    except Exception as e:
        print(f'获取浏览器实例失败！')
        # print(f'错误是：{e}')
        return webdriver.Chrome()


# 爬取某些设置下的所有学校
def get_page(url, driver):
    driver = driver
    driver.get(url)
    # 等待页面加载完毕，不然有可能找不到元素
    time.sleep(200)

    # 获取文件头
    # 定位到 tr 元素，这里使用 xpath 选择器
    tr_element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[3]/div[1]/div[2]/table/thead/tr')
    # 获取 tr 元素下的所有 th 元素
    th_elements = tr_element.find_elements(By.TAG_NAME, 'th')

    # 提取每个 th 元素的文本
    headers = []
    for th in th_elements:
        # 滑窗，滚动到定位元素
        driver.execute_script("arguments[0].scrollIntoView();", th)
        headers.append(th.text)
        # print(th.text)

    headers.pop(-1)
    sv_info.append(headers)
    error_sv_info.append(headers)

    header_len = len(headers)
    print(headers)
    print(f'设置的字段数共有：{header_len}')

    # 获取SV内容
    pages = []  # 一会用于判断是不是存在重复页

    # count = 1
    while True:
        # if count >= 3:  # 纯粹用来测试的
        #     break
        # else:
        #     count += 1

        # 判断是不是最后一页
        # 定位到 ul 元素，这里使用 xpath 选择器
        ul_element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[3]/div[2]/div/div/ul')

        # 获取ul元素下的所有直接子元素
        li_elements = ul_element.find_elements(By.XPATH, './li')

        page_num = 0  # 初始值，以防没有class_value == 'number active'的元素
        # 遍历所有li元素，获取每个元素的class属性值
        for li in li_elements:
            # 获取class属性值
            class_value = li.get_attribute('class')
            if class_value == 'number active':
                page_num = li.text

            # # 打印class属性值
            # print(f'"{class_value}"')

        # # 定位到 ul 下具有特定 class 的标签，假设 class 名称为 'my-class'
        # page_num = ul_element.find_element(By.CLASS_NAME, 'number active')

        # 提取标签的文本
        print(f'当前的页数是：{page_num}')
        if page_num in pages:
            print(f'当前页已经重复了！')
            break
        else:
            pages.append(page_num)

        # 获取当前页面每个SV的信息
        # 获取整个tbody
        tbody_element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[3]/div[1]/div[3]/table/tbody')

        # 定位到tbody下的所有tr元素
        tr_elements = tbody_element.find_elements(By.TAG_NAME, 'tr')

        # 遍历所有tr元素，提取每个tr下的文本
        for tr in tr_elements:
            current_sv = []
            # 获取tr下的所有直接子元素
            child_elements = tr.find_elements(By.XPATH, './*')
            # 获取每个子元素的文本并打印
            for child in child_elements:
                current_sv.append(child.text)
            # print(current_sv)
            if len(current_sv) == header_len:
                sv_info.append(current_sv)
                # print(f'字段数相符！')
            else:
                error_sv_info.append(current_sv)
                # print(f'字段数不相符！可能出问题了')

        # 如果是最后一页 4421
        if int(page_num) == 4421:
            # 可能已经到达最后一页
            print(f'当前的页数是：{page_num}，可能已经是最后一页了！')
            break

        # 点击下一页
        while True:
            try:
                button = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[3]/div[2]/div/div/button[2]')
                button.click()
                time.sleep(5)

                # 判断是否翻页
                # 定位到 ul 元素，这里使用 xpath 选择器
                ul_element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[3]/div[2]/div/div/ul')

                # 获取ul元素下的所有直接子元素
                li_elements = ul_element.find_elements(By.XPATH, './li')

                current_page = 0  # 初始值，以防没有class_value == 'number active'的元素
                # 遍历所有li元素，获取每个元素的class属性值
                for li in li_elements:
                    # 获取class属性值
                    class_value = li.get_attribute('class')
                    if class_value == 'number active':
                        current_page = li.text
                if int(current_page) == (int(page_num) + 1):
                    # 成功翻页
                    break
                else:
                    # 没有翻页，或者翻页出错，等待5S，然后再次尝试翻页操作
                    time.sleep(5)
                    continue
            except:
                continue


def export_info():
    # 输出excel
    # # 创建输出表格Excel：创建工作表
    # excel = openpyxl.Workbook()
    # # 创建sheet页：以demo为名字创建一个sheet页
    # sheet = excel.create_sheet('全部SV_info', 0)
    #
    # # 第一行第一列的单元格
    # for i in range(len(sv_info)):
    #     for k in range(len(sv_info[i])):
    #         cell = sheet.cell(row=i + 1, column=k + 1)
    #         # 单元格赋值
    #         cell.value = sv_info[i][k]
    #
    # # 保存excel文件
    # excel.save('sv_info.xlsx')

    # 输出csv
    # 打开文件，写入数据

    # 正常信息
    with open("sv_info.csv", mode='w', newline='', encoding='utf-8') as file:
        # 创建一个csv.writer对象
        csv_writer = csv.writer(file)

        # 写入二维列表的每一行
        for row in sv_info:
            csv_writer.writerow(row)

    # 异常信息
    with open("error_sv_info.csv", mode='w', newline='', encoding='utf-8') as file:
        # 创建一个csv.writer对象
        csv_writer = csv.writer(file)

        # 写入二维列表的每一行
        for row in error_sv_info:
            csv_writer.writerow(row)


def main():
    # 设置你想要搜索的问题
    web_url = 'https://www.biosino.org/pggsv/#/toolsTableBrowser'
    driver = get_driver()
    get_page(web_url, driver)
    export_info()

    # 休息一分钟，检查信息
    time.sleep(10)

    # 不关闭浏览器
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("t").key_up(Keys.CONTROL).perform()


if __name__ == "__main__":
    main()
