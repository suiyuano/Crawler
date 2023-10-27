# 好的，您可以使用Python中的Selenium库来模拟浏览器访问网站并复制所有内容，以下是代码示例：
# pythonCopy Code

import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

# 引入必要的库
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import re
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl


# 全局变量
# school_code = {}
school_info = []


def get_driver():
    try:
        # 设置浏览器驱动路径（这里使用Chrome浏览器驱动作为示例）
        driver_path = 'E:\软件\chromedriver\chromedriver_win32\chromedriver'

        # 选择谷歌浏览器
        s = Service(executable_path=driver_path)
        driver = webdriver.Chrome(service=s)
        return driver
    except Exception:
        print(f'获取浏览器实例失败！')
        return webdriver.Chrome()


# 得到登录的cookie
def login_cookie():
    driver = get_driver()
    driver.set_page_load_timeout(5)
    driver.set_script_timeout(5)
    LOGIN_URL = 'https://www.ewt360.com/'
    driver.get(LOGIN_URL)
    time.sleep(5)
    input("请登录后按 Enter")
    cookies = driver.get_cookies()
    jsonCookies = json.dumps(cookies)
    # 下面的文件位置需要自己改
    with open('cookie.txt', 'w') as f:
        f.write(jsonCookies)
    driver.quit()


# 再次登录
def login():
    driver.set_page_load_timeout(5)
    driver.set_script_timeout(5)
    LOGIN_URL = 'https://www.ewt360.com/'
    driver.get(LOGIN_URL)
    time.sleep(5)
    # 下面的文件位置需要自己改，与上面的改动一致
    f = open('cookie.txt')
    cookies = f.read()
    jsonCookies = json.loads(cookies)
    # print(jsonCookies)
    for co in jsonCookies:
        driver.add_cookie(co)
    # for co in cookie:
    #     driver.add_cookie(co)
    driver.refresh()
    # 打印网页title
    # print(driver.title)
    time.sleep(5)


"""
根据table的id属性和table中的某一个元素定位其在table中的位置
table包括表头，位置坐标都是从1开始算
tableId：table的id属性
queryContent：需要确定位置的内容
"""


def get_table_content(tableId):
    page_schools = []
    table_loc = (By.XPATH, tableId)
    # 按行查询表格的数据，取出的数据是一整行，按空格分隔每一列的数据
    table_tr_list = driver.find_element(*table_loc).find_elements(By.TAG_NAME, "tr")

    for tr in table_tr_list:
        # print(tr.text)
        # if tr.text == "院校名称 专业名称 本专科批次 首选科目 再选科目 收藏":
        #     # 是标题，直接略过
        #     continue
        current_school_info = (tr.text).split("\n")  # 以换行符拆分成若干个(个数与列的个数相同)一维列表

        page_schools.append(current_school_info)  # 将表格数据组成二维的列表
        school_info.append(current_school_info)  # 将当前表格存进总的信息池

    print(page_schools)

    # # 判断当前页数
    # try:
    #     page_num = driver.find_element(By.XPATH, '//*[@id="pagination"]/div[1]').text.split('：')[1].split(' ')[0]
    #     print(f'当前页数是：{page_num}')
    # except:
    #     text = driver.find_element(By.XPATH, '//*[@id="pagination"]/div[1]').text
    #     print(f'出错了！当前的text：{text}')
    #     print(page_schools)
    #
    #     message = 'break'
    #     return message
    #
    # if page_num != "3787/3787":
    #     message = 'continue'
    # else:
    #     message = 'break'
    #
    # print(page_schools)
    # return message


# 爬取某些设置下的所有学校
def get_schools(url):
    driver.get(url)
    # 等待页面加载完毕，不然有可能找不到元素
    time.sleep(5)

    # 先刷新input 文本框的内容
    input_element = driver.find_element(By.ID, 'rec-input1')  # 获取该输入框的ID
    input_element.clear()  # 清楚该输入框中的原本内容
    input_element.send_keys("1", Keys.ENTER)  # 向该输入框中添加位次排名: "1“
    time.sleep(10)

    input_element = driver.find_element(By.XPATH,
                                        '//*[@id="root"]/section/section/section/div[1]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div/input[2]')  # 获取该输入框的XPATH
    input_element.clear()  # 清楚该输入框中的原本内容
    input_element.send_keys("1000000", Keys.ENTER)  # 向该输入框中添加位次排名: "1000000“
    # input_element.send_keys("100", Keys.ENTER)  # 向该输入框中添加位次排名: "100“ test
    time.sleep(10)



    # count=0
    while True:
        # if count>=3:
        #     break
        # else:
        #     count+=1

        # message = get_table_content('//*[@id="old-home-data-list"]/div/div/table/tbody')
        # if message == "break":
        #     print("开始重复了！请检查！")
        #     break
        # elif message == "continue":
        #     pass
        # else:
        #     print("出现未知错误！请检查！")
        #     break

        get_table_content('//*[@id="old-home-data-list"]/div/div/table/tbody')

        page = driver.find_element(By.XPATH,
                                   '//*[@id="root"]/section/section/section/div[1]/div/div[2]/div/div[2]/div[2]/div[2]/ul')

        list = page.find_elements(By.XPATH, 'li')
        # len(list)  # 计算有多少个a
        last_li = list[-1]  # 用列表标识符取最后一个li
        last_li_title = last_li.get_attribute("title")
        if last_li_title == "Next Page":
            # 是下一页

            # 判断是否是最后一页
            sig = last_li.get_attribute("aria-disabled")
            if sig == "false":
                # 不是最后一页
                print(f'最后一个元素的title是：{last_li_title}')
                next_page = last_li

            elif sig == 'true':
                # 是最后一页
                break
            else:
                print('翻页时候出错了！请检查！')
                break

        else:
            print(f'最后一个元素的title是：{last_li_title}')
            break

        # # 获取用于判断是否是最后一页的属性
        # is_next_url = next_page.get_attribute("aria-disabled")
        # print(is_next_url)

        try:
            # next_page.click() #可能会报错，原因可能是由悬浮元素遮挡
            driver.execute_script("(arguments[0]).click()", next_page)
            time.sleep(5)
        except:
            print('已到最后一页！ Task finished！')
            break

    # number = int(re.search('[0-9]+', number_text).group())
    # driver.find_element_by_partial_link_text('查看全部').click()
    # for k in range(number):
    #     xpath = '/html/body/div[1]/div/main/div/div[2]/div[1]/div/div[2]/div/div/div/div[2]/div/div[{}]/div/div[2]/div[1]/span'.format(
    #         k + 1)
    #     element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
    #     answer = element.text
    #     # 下面的文件位置需要自己改，保存到你想要的位置
    #     file = open('C:/Users/zxw/Desktop/修身/与自己/数据分析/数据分析/爬虫/回答/answer{}.txt'.format(k + 1), 'w',
    #                 encoding='utf-8')
    #     file.write(answer)
    #     file.close()
    #     print('answer ' + str(k + 1) + ' collected!')
    #     time.sleep(1)
    #     js = "window.scrollTo(0,document.body.scrollHeight)"
    #     driver.execute_script(js)
    #     time.sleep(1)


def export_info():
    # 创建输出表格Excel：创建工作表
    excel = openpyxl.Workbook()
    # 创建sheet页：以demo为名字创建一个sheet页
    sheet = excel.create_sheet('本科一批', 0)

    # 第一行第一列的单元格
    for i in range(len(school_info)):
        for k in range(len(school_info[i])):
            cell = sheet.cell(row=i + 1, column=k + 1)
            # 单元格赋值
            cell.value = school_info[i][k]

    # 保存excel文件
    excel.save('school_info.xlsx')


if __name__ == "__main__":
    # 设置你想要搜索的问题
    web_url = 'https://www.ewt360.com/career/careerlibs/index.html#/volunteerold?queryType=2'
    login_cookie()
    driver = get_driver()
    login()
    get_schools(web_url)
    export_info()

    # 休息一分钟，检查信息
    time.sleep(10)

    # 不关闭浏览器
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("t").key_up(Keys.CONTROL).perform()
